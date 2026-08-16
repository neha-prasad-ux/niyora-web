#!/usr/bin/env python3
"""Export the Research sheet of niyora-research-table.xlsx to src/data/research.json.

The spreadsheet stays the source of truth. Run this after editing it:

    python3 scripts/research-export.py

The internal columns (Use it for, Gap / to add) are never exported. Clustering
lives in src/pages/research/_clusters.ts, not here.
"""

import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SHEET = ROOT.parent / "niyora-research-table.xlsx"
OUT = ROOT / "src/data/research.json"

URL = re.compile(r"https?://\S+")


def main() -> int:
    if not SHEET.exists():
        sys.exit(f"spreadsheet not found: {SHEET}")

    ws = openpyxl.load_workbook(SHEET, data_only=True)["Research"]
    header = [str(c.value).strip() if c.value else "" for c in ws[1]]
    idx = {name: i for i, name in enumerate(header)}

    rows, unlinked = [], []
    for values in ws.iter_rows(min_row=2, values_only=True):
        def get(name: str) -> str:
            i = idx.get(name)
            return str(values[i]).strip() if i is not None and values[i] else ""

        if not get("Topic"):
            continue

        source = get("Source")
        # A link pasted into Source counts; the Link column just overrides it.
        found = URL.search(source)
        link = get("Link") or (found.group(0) if found else "")

        row = {
            "n": int(get("#")),
            "topic": get("Topic"),
            "tags": [t.strip() for t in get("Tags").split(",") if t.strip()],
            "finding": get("Finding (the research)"),
            "grade": get("Grade").lower(),
            "plain": get("Plain line (what she reads)"),
            "source": URL.sub("", source).strip(" -—"),
            "link": link,
            "status": get("Status").lower(),
        }
        if not row["link"]:
            unlinked.append(f'  {row["n"]}. {row["topic"]} ({row["source"]})')
        rows.append(row)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(rows, indent=2, ensure_ascii=False) + "\n")
    print(f"wrote {len(rows)} rows to {OUT.relative_to(ROOT)}")

    if unlinked:
        print(f"\n{len(unlinked)} rows have no paper link (they publish without one):")
        print("\n".join(unlinked))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
